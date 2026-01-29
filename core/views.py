from django.shortcuts import render, redirect, get_object_or_404
import datetime
from django.core.paginator import Paginator
from .models import Student, Teacher, TeacherAssignment, Attendance, Grade, Enrollment, FeeStructureItem, Payment, Course
from django.db import models
from .forms import StudentForm, TeacherForm, CourseForm, TeacherAssignmentForm, EnrollmentForm, AttendanceForm, GradeForm, PaymentForm
import openpyxl
from django.contrib import messages
from .utils import log_action
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from io import BytesIO
from django.contrib.auth.decorators import login_required
from .decorators import role_required

@login_required
def home(request):
    return render(request, 'home.html')

@login_required
@role_required(['Admin', 'Principal', 'Clerk'])
def student_bulk_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('student_bulk_import')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            students_to_create = []
            errors = []

            # Skip header
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue # skip empty rows

                name, dob, gender, class_grade, g_name, g_contact, g_email, address, medical = row[:9]

                # Basic validation
                if not all([name, dob, gender, class_grade, g_name, g_contact]):
                    errors.append(f"Row {i}: Missing required fields.")
                    continue

                try:
                    # Parse DOB if it's a string
                    if isinstance(dob, str):
                        dob = datetime.datetime.strptime(dob, '%Y-%m-%d').date()

                    student = Student(
                        full_name=name,
                        date_of_birth=dob,
                        gender=gender,
                        class_grade=class_grade,
                        guardian_name=g_name,
                        guardian_contact=g_contact,
                        guardian_email=g_email,
                        home_address=address or "",
                        medical_conditions=medical or "",
                        created_by=request.user
                    )
                    student.full_clean()
                    students_to_create.append(student)
                except Exception as e:
                    errors.append(f"Row {i}: {str(e)}")

            if errors:
                for error in errors:
                    messages.error(request, error)
            else:
                for student in students_to_create:
                    student.save()
                log_action(request.user, "Bulk Student Import", f"Imported {len(students_to_create)} students", request)
                messages.success(request, f"Successfully imported {len(students_to_create)} students.")
                return redirect('student_list')

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

    return render(request, 'student_bulk_import.html')

@login_required
@role_required(['Admin', 'Principal', 'Clerk'])
def student_registration(request):
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()
            log_action(request.user, "Student Registered", f"Registered student {student.full_name} ({student.student_id})", request)
            return redirect('home')  # Redirect to a success page or student list
    else:
        form = StudentForm()
    return render(request, 'student_registration.html', {'form': form})

@login_required
@role_required(['Admin', 'Principal', 'Clerk'])
def export_students_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="students.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ['ID', 'Full Name', 'DOB', 'Gender', 'Class', 'Guardian Name', 'Guardian Contact', 'Status']
    ws.append(headers)

    students = Student.objects.filter(is_deleted=False)
    for s in students:
        ws.append([s.student_id, s.full_name, s.date_of_birth, s.gender, s.class_grade, s.guardian_name, s.guardian_contact, s.status])

    wb.save(response)
    return response

@login_required
@role_required(['Admin', 'Principal', 'Clerk'])
def student_list(request):
    students_query = Student.objects.filter(is_deleted=False).order_by('full_name')

    # Search and filter
    query = request.GET.get('q')
    class_filter = request.GET.get('class_grade')
    status_filter = request.GET.get('status')

    if query:
        students_query = students_query.filter(
            models.Q(full_name__icontains=query) |
            models.Q(student_id__icontains=query) |
            models.Q(guardian_contact__icontains=query)
        )

    if class_filter:
        students_query = students_query.filter(class_grade=class_filter)

    if status_filter:
        students_query = students_query.filter(status=status_filter)

    paginator = Paginator(students_query, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'class_choices': Student.CLASS_CHOICES,
        'status_choices': Student.STATUS_CHOICES,
        'query': query,
        'class_filter': class_filter,
        'status_filter': status_filter,
    }

    return render(request, 'student_list.html', context)

@login_required
@role_required(['Admin', 'Principal', 'Clerk', 'Teacher'])
def student_profile(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    # Attendance summary
    attendance_records = Attendance.objects.filter(student=student)
    total_days = attendance_records.count()
    present_days = attendance_records.filter(status='Present').count()
    attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0

    # Grade summary (average of recent grades)
    recent_grades = Grade.objects.filter(student=student).order_by('-recorded_at')[:10]
    avg_grade = recent_grades.aggregate(models.Avg('score'))['score__avg'] or 0

    # Fee summary
    # Total fees for the student's class
    fee_items = FeeStructureItem.objects.filter(fee_structure__class_grade=student.class_grade)
    total_fees = fee_items.aggregate(models.Sum('amount'))['amount__sum'] or 0

    # Payments made
    paid_fees = Payment.objects.filter(student=student).aggregate(models.Sum('amount'))['amount__sum'] or 0
    balance = total_fees - paid_fees

    context = {
        'student': student,
        'attendance_percentage': attendance_percentage,
        'present_days': present_days,
        'total_days': total_days,
        'avg_grade': avg_grade,
        'total_fees': total_fees,
        'paid_fees': paid_fees,
        'balance': balance,
    }
    return render(request, 'student_profile.html', context)

@login_required
@role_required(['Admin', 'Principal', 'Clerk'])
def student_update(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            student = form.save(commit=False)
            student.modified_by = request.user
            student.save()
            log_action(request.user, "Student Updated", f"Updated student {student.full_name} ({student.student_id})", request)
            return redirect('student_profile', student_id=student.id)
    else:
        form = StudentForm(instance=student)
    return render(request, 'student_update.html', {'form': form, 'student': student})

@login_required
@role_required(['Admin', 'Principal'])
def student_delete(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        reason = request.POST.get('reason')
        if reason:
            student.is_deleted = True
            student.deletion_reason = reason
            student.modified_by = request.user
            student.save()
            log_action(request.user, "Student Deleted", f"Deleted student {student.full_name} ({student.student_id}). Reason: {reason}", request)
            return redirect('student_list')
    return render(request, 'student_delete.html', {'student': student})

@login_required
def teacher_timetable(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    assignments = TeacherAssignment.objects.filter(teacher=teacher)

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    slots = [
        '08:00 - 09:00', '09:00 - 10:00', '10:00 - 11:00', '11:00 - 12:00',
        '12:00 - 13:00', '14:00 - 15:00', '15:00 - 16:00', '16:00 - 17:00'
    ]

    # Organize assignments into a grid
    timetable = {slot: {day: None for day in days} for slot in slots}
    for a in assignments:
        if a.time_slot in timetable and a.day_of_week in timetable[a.time_slot]:
            timetable[a.time_slot][a.day_of_week] = a

    return render(request, 'teacher_timetable.html', {
        'teacher': teacher,
        'timetable': timetable,
        'days': days,
        'slots': slots
    })

@login_required
@role_required(['Admin', 'Principal'])
def teacher_list(request):
    teachers = Teacher.objects.filter(is_deleted=False).order_by('full_name')
    return render(request, 'teacher_list.html', {'teachers': teachers})

@login_required
@role_required(['Admin', 'Principal'])
def teacher_registration(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            teacher = form.save(commit=False)
            teacher.created_by = request.user
            teacher.save()
            form.save_m2m()
            log_action(request.user, "Teacher Registered", f"Registered teacher {teacher.full_name} ({teacher.teacher_id})", request)
            return redirect('teacher_profile', teacher_id=teacher.id)
    else:
        form = TeacherForm()
    return render(request, 'teacher_registration.html', {'form': form})

@login_required
@role_required(['Admin', 'Principal'])
def teacher_profile(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    return render(request, 'teacher_profile.html', {'teacher': teacher})

@login_required
@role_required(['Admin', 'Principal'])
def course_list(request):
    courses = Course.objects.all().order_by('grade_level', 'course_name')
    return render(request, 'course_list.html', {'courses': courses})

@login_required
@role_required(['Admin', 'Principal'])
def course_creation(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save()
            log_action(request.user, "Course Created", f"Created course {course.course_name} ({course.course_code})", request)
            return redirect('course_list')
    else:
        form = CourseForm()
    return render(request, 'course_creation.html', {'form': form})

@login_required
@role_required(['Admin', 'Principal', 'Teacher', 'Clerk'])
def teacher_assignment(request):
    if request.method == 'POST':
        form = TeacherAssignmentForm(request.POST)
        if form.is_valid():
            assignment = form.save()
            log_action(request.user, "Teacher Assigned", f"Assigned {assignment.teacher.full_name} to {assignment.course.course_name} for {assignment.class_grade}", request)
            return redirect('home') # Redirect to assignment list page later
    else:
        form = TeacherAssignmentForm()
    return render(request, 'teacher_assignment.html', {'form': form})

@login_required
@role_required(['Admin', 'Principal', 'Teacher', 'Clerk'])
def bulk_enrollment(request):
    if request.method == 'POST':
        class_grade = request.POST.get('class_grade')
        term = request.POST.get('term')
        academic_year = request.POST.get('academic_year')
        student_ids = request.POST.getlist('students')
        course_ids = request.POST.getlist('courses')

        if class_grade and term and academic_year and student_ids and course_ids:
            count = 0
            for student_id in student_ids:
                for course_id in course_ids:
                    Enrollment.objects.get_or_create(
                        student_id=student_id,
                        course_id=course_id,
                        term=term,
                        academic_year=academic_year
                    )
                    count += 1
            log_action(request.user, "Bulk Enrollment", f"Enrolled students in {count} course slots for {class_grade}, {term} {academic_year}", request)
            messages.success(request, f"Successfully created {count} enrollment records.")
            return redirect('home')
        else:
            messages.error(request, "Please select all required fields.")

    students = Student.objects.filter(is_deleted=False)
    courses = Course.objects.filter(status='Active')

    context = {
        'students': students,
        'courses': courses,
        'class_choices': Student.CLASS_CHOICES,
        'term_choices': TeacherAssignment.TERM_CHOICES,
    }
    return render(request, 'bulk_enrollment.html', context)

@login_required
@role_required(['Admin', 'Principal', 'Teacher', 'Clerk'])
def enrollment(request):
    if request.method == 'POST':
        form = EnrollmentForm(request.POST)
        if form.is_valid():
            enrollment = form.save()
            log_action(request.user, "Student Enrolled", f"Enrolled student {enrollment.student.full_name} in {enrollment.course.course_name}", request)
            return redirect('home') # Redirect to enrollment list page later
    else:
        form = EnrollmentForm()
    return render(request, 'enrollment.html', {'form': form})

@login_required
@role_required(['Admin', 'Principal', 'Teacher'])
def bulk_attendance(request):
    courses = Course.objects.filter(status='Active')
    selected_course = None
    students = []

    date = request.GET.get('date') or datetime.date.today().strftime('%Y-%m-%d')
    course_id = request.GET.get('course')

    if course_id:
        selected_course = get_object_or_404(Course, id=course_id)
        # Find students enrolled in this course for the current term/year
        # To keep it simple, we'll just get all students enrolled in this course
        enrollments = Enrollment.objects.filter(course=selected_course).select_related('student')
        students = [e.student for e in enrollments]

    if request.method == 'POST':
        date = request.POST.get('date')
        course_id = request.POST.get('course_id')
        selected_course = get_object_or_404(Course, id=course_id)

        for key, value in request.POST.items():
            if key.startswith('attendance_'):
                student_id = key.split('_')[1]
                Attendance.objects.update_or_create(
                    student_id=student_id,
                    course=selected_course,
                    date=date,
                    defaults={'status': value}
                )
        log_action(request.user, "Bulk Attendance", f"Recorded attendance for {selected_course.course_name} on {date}", request)
        messages.success(request, "Attendance recorded successfully.")
        return redirect('home')

    return render(request, 'bulk_attendance.html', {
        'courses': courses,
        'selected_course': selected_course,
        'students': students,
        'date': date,
        'status_choices': Attendance.STATUS_CHOICES
    })

@login_required
@role_required(['Admin', 'Principal', 'Teacher'])
def student_attendance_report(request):
    students = Student.objects.filter(is_deleted=False)
    selected_student = None
    records = []
    summary = {}

    student_id = request.GET.get('student')
    if student_id:
        selected_student = get_object_or_404(Student, id=student_id)
        records = Attendance.objects.filter(student=selected_student).order_by('-date')

        total = records.count()
        present = records.filter(status='Present').count()
        absent = records.filter(status='Absent').count()
        late = records.filter(status='Late').count()
        excused = records.filter(status='Excused').count()

        summary = {
            'total': total,
            'present': present,
            'absent': absent,
            'late': late,
            'excused': excused,
            'percentage': (present / total * 100) if total > 0 else 0
        }

    return render(request, 'student_attendance_report.html', {
        'students': students,
        'selected_student': selected_student,
        'records': records,
        'summary': summary
    })

@login_required
@role_required(['Admin', 'Principal', 'Teacher'])
def class_attendance_report(request):
    class_grade = request.GET.get('class_grade')
    summary = []

    if class_grade:
        students = Student.objects.filter(class_grade=class_grade, is_deleted=False)
        for student in students:
            records = Attendance.objects.filter(student=student)
            total = records.count()
            present = records.filter(status='Present').count()
            summary.append({
                'student': student,
                'total': total,
                'present': present,
                'percentage': (present / total * 100) if total > 0 else 0
            })

    return render(request, 'class_attendance_report.html', {
        'summary': summary,
        'class_choices': Student.CLASS_CHOICES,
        'selected_class': class_grade
    })

@login_required
@role_required(['Admin', 'Principal', 'Teacher'])
def attendance(request):
    if request.method == 'POST':
        form = AttendanceForm(request.POST)
        if form.is_valid():
            att = form.save()
            log_action(request.user, "Attendance Recorded", f"Recorded attendance for {att.student.full_name} on {att.date}", request)
            return redirect('home') # Redirect to attendance list page later
    else:
        form = AttendanceForm()
    return render(request, 'attendance.html', {'form': form})

@login_required
@role_required(['Admin', 'Accountant'])
def payment_list(request):
    payments = Payment.objects.all().order_by('-date')
    total_collected = payments.aggregate(models.Sum('amount'))['amount__sum'] or 0
    return render(request, 'payment_list.html', {
        'payments': payments,
        'total_collected': total_collected
    })

@login_required
@role_required(['Admin', 'Accountant'])
def record_payment(request):
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.recorded_by = request.user
            payment.save()
            log_action(request.user, "Payment Recorded", f"Recorded payment of {payment.amount} for {payment.student.full_name}. Receipt: {payment.receipt_number}", request)
            messages.success(request, "Payment recorded successfully.")
            return redirect('payment_list')
    else:
        form = PaymentForm()
    return render(request, 'record_payment.html', {'form': form})

@login_required
def generate_receipt(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)

    buffer = BytesIO()
    p = canvas.Canvas(buffer)

    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, "OFFLINE SCHOOL MANAGEMENT SYSTEM")
    p.setFont("Helvetica", 12)
    p.drawString(100, 780, "Official Payment Receipt")
    p.line(100, 775, 500, 775)

    p.drawString(100, 750, f"Receipt No: {payment.receipt_number}")
    p.drawString(100, 730, f"Date: {payment.date}")

    p.drawString(100, 700, f"Student Name: {payment.student.full_name}")
    p.drawString(100, 680, f"Student ID: {payment.student.student_id}")
    p.drawString(100, 660, f"Class: {payment.student.class_grade}")

    p.setFont("Helvetica-Bold", 14)
    p.drawString(100, 620, f"Amount Paid: {payment.amount} UGX")

    p.setFont("Helvetica", 10)
    p.drawString(100, 580, f"Recorded by: {payment.recorded_by.username}")
    p.drawString(100, 560, "Thank you for your payment.")

    p.showPage()
    p.save()

    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

@login_required
@role_required(['Admin', 'Principal', 'Teacher'])
def record_grade(request):
    if request.method == 'POST':
        form = GradeForm(request.POST)
        if form.is_valid():
            grade = form.save(commit=False)
            grade.recorded_by = request.user
            grade.save()
            log_action(request.user, "Grade Recorded", f"Recorded grade {grade.score} for {grade.student.full_name} on {grade.assessment.assessment_name}", request)
            return redirect('home') # Redirect to grade list page later
    else:
        form = GradeForm()
    return render(request, 'record_grade.html', {'form': form})
